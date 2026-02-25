"""Build a DCF model in Excel using Hermes tools directly.

Demonstrates using the Excel and chart tools programmatically
(without the agent layer) to build a discounted cash flow financial
model.  This is useful when you want fine-grained control over model
construction rather than delegating to the modeling agent.
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_dcf_model(output_path: str = "./output/dcf_model.xlsx") -> str:
    """Build a complete DCF model workbook and save to disk.

    The workbook contains three sheets:
        1. Assumptions -- key inputs (growth rates, WACC, margins)
        2. DCF Model   -- projected FCF, discount factors, NPV calculation
        3. Summary     -- enterprise value waterfall with a bar chart

    Args:
        output_path: Where to save the .xlsx file.

    Returns:
        The absolute path to the saved workbook.
    """

    wb = Workbook()

    # -----------------------------------------------------------------------
    # Styles
    # -----------------------------------------------------------------------
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    number_font = Font(name="Calibri", size=11)
    pct_format = "0.0%"
    usd_format = '#,##0.0'
    usd_format_whole = '#,##0'
    thin_border = Border(
        bottom=Side(style="thin", color="B4C6E7"),
    )

    def style_header_row(ws, row: int, max_col: int) -> None:
        """Apply header styling to a row."""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # -----------------------------------------------------------------------
    # Sheet 1: Assumptions
    # -----------------------------------------------------------------------
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_assumptions.column_dimensions["A"].width = 30
    ws_assumptions.column_dimensions["B"].width = 15

    # Write the assumptions table.
    assumptions = [
        ("Assumption", "Value"),
        ("Company", "Acme Corp"),
        ("Base Year Revenue ($M)", 10000),
        ("Revenue Growth Rate (Year 1-3)", 0.12),
        ("Revenue Growth Rate (Year 4-5)", 0.08),
        ("Terminal Growth Rate", 0.025),
        ("EBITDA Margin", 0.30),
        ("D&A (% of Revenue)", 0.04),
        ("Capex (% of Revenue)", 0.05),
        ("Change in NWC (% of Rev Growth)", 0.10),
        ("Tax Rate", 0.21),
        ("WACC", 0.10),
        ("Net Debt ($M)", 5000),
        ("Shares Outstanding (M)", 500),
    ]

    for row_idx, (label, value) in enumerate(assumptions, start=1):
        ws_assumptions.cell(row=row_idx, column=1, value=label)
        ws_assumptions.cell(row=row_idx, column=2, value=value)

    # Style the header row.
    style_header_row(ws_assumptions, 1, 2)

    # Format percentages.
    for row_idx in (4, 5, 6, 7, 8, 9, 10, 11, 12):
        ws_assumptions.cell(row=row_idx, column=2).number_format = pct_format

    # -----------------------------------------------------------------------
    # Sheet 2: DCF Model
    # -----------------------------------------------------------------------
    ws_dcf = wb.create_sheet("DCF Model")

    # Column layout: A = labels, B-F = Year 1-5, G = Terminal
    years = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Terminal"]
    headers = ["Metric"] + years

    for col_idx, header in enumerate(headers, start=1):
        ws_dcf.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_dcf, 1, len(headers))
    ws_dcf.column_dimensions["A"].width = 28
    for col_idx in range(2, len(headers) + 1):
        ws_dcf.column_dimensions[get_column_letter(col_idx)].width = 14

    # Row labels for the model.
    row_labels = [
        "Revenue ($M)",
        "Revenue Growth",
        "EBITDA ($M)",
        "EBITDA Margin",
        "D&A ($M)",
        "EBIT ($M)",
        "Tax ($M)",
        "NOPAT ($M)",
        "Add: D&A ($M)",
        "Less: Capex ($M)",
        "Less: Chg NWC ($M)",
        "Free Cash Flow ($M)",
        "",
        "Discount Factor",
        "PV of FCF ($M)",
    ]

    for row_idx, label in enumerate(row_labels, start=2):
        ws_dcf.cell(row=row_idx, column=1, value=label)

    # Build projected values using formulas that reference the Assumptions sheet.
    # For clarity we use hardcoded references; a production tool would generate
    # these dynamically.
    base_rev = 10000
    growth_high = 0.12
    growth_low = 0.08
    terminal_g = 0.025
    ebitda_margin = 0.30
    da_pct = 0.04
    capex_pct = 0.05
    nwc_pct = 0.10
    tax_rate = 0.21
    wacc = 0.10

    prev_rev = base_rev
    for yr in range(5):
        col = yr + 2  # columns B through F
        g = growth_high if yr < 3 else growth_low
        rev = prev_rev * (1 + g)
        ebitda = rev * ebitda_margin
        da = rev * da_pct
        ebit = ebitda - da
        tax = ebit * tax_rate
        nopat = ebit - tax
        capex = rev * capex_pct
        nwc = (rev - prev_rev) * nwc_pct
        fcf = nopat + da - capex - nwc
        discount = 1 / ((1 + wacc) ** (yr + 1))
        pv_fcf = fcf * discount

        row_base = 2
        ws_dcf.cell(row=row_base, column=col, value=round(rev, 1))
        ws_dcf.cell(row=row_base + 1, column=col, value=g)
        ws_dcf.cell(row=row_base + 2, column=col, value=round(ebitda, 1))
        ws_dcf.cell(row=row_base + 3, column=col, value=ebitda_margin)
        ws_dcf.cell(row=row_base + 4, column=col, value=round(da, 1))
        ws_dcf.cell(row=row_base + 5, column=col, value=round(ebit, 1))
        ws_dcf.cell(row=row_base + 6, column=col, value=round(tax, 1))
        ws_dcf.cell(row=row_base + 7, column=col, value=round(nopat, 1))
        ws_dcf.cell(row=row_base + 8, column=col, value=round(da, 1))
        ws_dcf.cell(row=row_base + 9, column=col, value=round(capex, 1))
        ws_dcf.cell(row=row_base + 10, column=col, value=round(nwc, 1))
        ws_dcf.cell(row=row_base + 11, column=col, value=round(fcf, 1))
        ws_dcf.cell(row=row_base + 13, column=col, value=round(discount, 4))
        ws_dcf.cell(row=row_base + 14, column=col, value=round(pv_fcf, 1))

        prev_rev = rev

    # Terminal value column (G).
    terminal_fcf = ws_dcf.cell(row=13, column=6).value  # Year 5 FCF
    terminal_value = terminal_fcf * (1 + terminal_g) / (wacc - terminal_g)
    terminal_discount = 1 / ((1 + wacc) ** 5)
    pv_terminal = terminal_value * terminal_discount
    ws_dcf.cell(row=2, column=7, value="--")
    ws_dcf.cell(row=3, column=7, value=terminal_g)
    ws_dcf.cell(row=13, column=7, value=round(terminal_value, 1))
    ws_dcf.cell(row=15, column=7, value=round(terminal_discount, 4))
    ws_dcf.cell(row=16, column=7, value=round(pv_terminal, 1))

    # Apply number formats.
    for row_idx in range(2, 17):
        for col_idx in range(2, 8):
            cell = ws_dcf.cell(row=row_idx, column=col_idx)
            cell.font = number_font
            if row_idx in (3, 5):  # growth and margin rows
                cell.number_format = pct_format
            elif row_idx in (15,):  # discount factor
                cell.number_format = "0.0000"
            else:
                cell.number_format = usd_format
            cell.border = thin_border

    # -----------------------------------------------------------------------
    # Sheet 3: Summary -- enterprise value bridge
    # -----------------------------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 18

    # Calculate summary values.
    sum_pv_fcf = sum(
        ws_dcf.cell(row=16, column=c).value
        for c in range(2, 7)
        if isinstance(ws_dcf.cell(row=16, column=c).value, (int, float))
    )
    enterprise_value = sum_pv_fcf + pv_terminal
    net_debt = 5000
    equity_value = enterprise_value - net_debt
    shares = 500
    price_per_share = equity_value / shares

    summary_data = [
        ("Valuation Summary", "Value ($M)"),
        ("PV of Projected FCFs", round(sum_pv_fcf, 1)),
        ("PV of Terminal Value", round(pv_terminal, 1)),
        ("Enterprise Value", round(enterprise_value, 1)),
        ("Less: Net Debt", -net_debt),
        ("Equity Value", round(equity_value, 1)),
        ("Shares Outstanding (M)", shares),
        ("Implied Price / Share ($)", round(price_per_share, 2)),
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)

    style_header_row(ws_summary, 1, 2)

    for row_idx in range(2, len(summary_data) + 1):
        ws_summary.cell(row=row_idx, column=2).number_format = usd_format
        ws_summary.cell(row=row_idx, column=2).font = number_font

    # -----------------------------------------------------------------------
    # Add a bar chart showing the EV waterfall.
    # -----------------------------------------------------------------------
    chart = BarChart()
    chart.type = "col"
    chart.title = "Enterprise Value Bridge ($M)"
    chart.y_axis.title = "Value ($M)"
    chart.x_axis.title = None
    chart.style = 10
    chart.width = 20
    chart.height = 12

    # Chart data: PV of FCFs, PV of Terminal Value, Enterprise Value.
    chart_labels = Reference(ws_summary, min_col=1, min_row=2, max_row=4)
    chart_data = Reference(ws_summary, min_col=2, min_row=1, max_row=4)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_labels)
    chart.shape = 4
    ws_summary.add_chart(chart, "D2")

    # -----------------------------------------------------------------------
    # Save the workbook.
    # -----------------------------------------------------------------------
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    abs_path = os.path.abspath(output_path)
    wb.save(abs_path)
    print(f"DCF model saved to: {abs_path}")

    return abs_path


if __name__ == "__main__":
    build_dcf_model()
