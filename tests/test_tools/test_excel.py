"""Tests for Excel workbook tools.

All tests create real .xlsx files in temporary directories and verify
their contents using openpyxl.  No mocking of Excel functionality.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Helpers -- these mirror the operations that hermes/tools/excel.py will
# expose.  We test the underlying openpyxl behaviour directly so that the
# tests remain valid regardless of the tool wrapper implementation.
# ---------------------------------------------------------------------------


def create_workbook_on_disk(path: Path, sheet_name: str = "Sheet1") -> Workbook:
    """Create a new workbook with one sheet and save it to *path*."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(str(path))
    return wb


class TestCreateWorkbook:
    """Test workbook creation and basic file operations."""

    def test_creates_valid_xlsx_file(self, tmp_output_dir: Path) -> None:
        """A newly created workbook should be a valid .xlsx loadable by openpyxl."""
        path = tmp_output_dir / "test_create.xlsx"
        create_workbook_on_disk(path)

        # Verify the file exists and can be loaded.
        assert path.exists()
        wb = load_workbook(str(path))
        assert len(wb.sheetnames) >= 1
        wb.close()

    def test_custom_sheet_name(self, tmp_output_dir: Path) -> None:
        """The active sheet should use the provided name."""
        path = tmp_output_dir / "test_sheet_name.xlsx"
        create_workbook_on_disk(path, sheet_name="Financials")

        wb = load_workbook(str(path))
        assert "Financials" in wb.sheetnames
        wb.close()


class TestWriteAndReadCells:
    """Test writing values to cells and reading them back."""

    def test_write_string_and_read_back(self, tmp_output_dir: Path) -> None:
        """String values should round-trip through save/load."""
        path = tmp_output_dir / "test_strings.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Revenue"
        ws["B1"] = "Expenses"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2["A1"].value == "Revenue"
        assert ws2["B1"].value == "Expenses"
        wb2.close()

    def test_write_numbers_and_read_back(self, tmp_output_dir: Path) -> None:
        """Numeric values should preserve their type through save/load."""
        path = tmp_output_dir / "test_numbers.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 42
        ws["A2"] = 3.14159
        ws["A3"] = -1000.50
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2["A1"].value == 42
        assert abs(ws2["A2"].value - 3.14159) < 1e-5
        assert ws2["A3"].value == -1000.50
        wb2.close()

    def test_write_grid_of_values(self, tmp_output_dir: Path) -> None:
        """A grid of values written by row/column index should all persist."""
        path = tmp_output_dir / "test_grid.xlsx"
        wb = Workbook()
        ws = wb.active

        # Write a 5x3 grid.
        expected = {}
        for row in range(1, 6):
            for col in range(1, 4):
                value = row * 100 + col
                ws.cell(row=row, column=col, value=value)
                expected[(row, col)] = value
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        for (row, col), value in expected.items():
            assert ws2.cell(row=row, column=col).value == value
        wb2.close()


class TestFormulas:
    """Test that Excel formulas are written correctly."""

    def test_sum_formula(self, tmp_output_dir: Path) -> None:
        """A SUM formula should be preserved as a formula string."""
        path = tmp_output_dir / "test_formulas.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        ws["A2"] = 200
        ws["A3"] = 300
        ws["A4"] = "=SUM(A1:A3)"
        wb.save(str(path))

        # Load without evaluating formulas (data_only=False is the default).
        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2["A4"].value == "=SUM(A1:A3)"
        wb2.close()

    def test_percentage_formula(self, tmp_output_dir: Path) -> None:
        """Division formulas should be written verbatim."""
        path = tmp_output_dir / "test_pct_formula.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 50
        ws["A2"] = 200
        ws["A3"] = "=A1/A2"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2["A3"].value == "=A1/A2"
        wb2.close()

    def test_cross_sheet_reference(self, tmp_output_dir: Path) -> None:
        """Formulas referencing other sheets should preserve the sheet name."""
        path = tmp_output_dir / "test_cross_sheet.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Inputs"
        ws1["A1"] = 0.10  # WACC

        ws2 = wb.create_sheet("DCF")
        ws2["A1"] = "=Inputs!A1"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        assert wb2["DCF"]["A1"].value == "=Inputs!A1"
        wb2.close()


class TestFormatting:
    """Test cell formatting (fonts, fills, number formats)."""

    def test_bold_font(self, tmp_output_dir: Path) -> None:
        """Bold font should persist through save/load."""
        path = tmp_output_dir / "test_bold.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A1"].font = Font(bold=True, size=14, name="Calibri")
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        cell = wb2.active["A1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
        wb2.close()

    def test_fill_colour(self, tmp_output_dir: Path) -> None:
        """Background fill colour should persist."""
        path = tmp_output_dir / "test_fill.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Highlighted"
        ws["A1"].fill = PatternFill(
            start_color="FF2F5496", end_color="FF2F5496", fill_type="solid"
        )
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        cell = wb2.active["A1"]
        assert cell.fill.start_color.rgb == "FF2F5496"
        wb2.close()

    def test_number_format_percentage(self, tmp_output_dir: Path) -> None:
        """Percentage number format should be preserved."""
        path = tmp_output_dir / "test_pct_format.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 0.125
        ws["A1"].number_format = "0.0%"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        assert wb2.active["A1"].number_format == "0.0%"
        assert wb2.active["A1"].value == 0.125
        wb2.close()

    def test_number_format_currency(self, tmp_output_dir: Path) -> None:
        """Currency number format should be preserved."""
        path = tmp_output_dir / "test_currency_format.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1234567.89
        ws["A1"].number_format = '$#,##0.00'
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        assert wb2.active["A1"].number_format == '$#,##0.00'
        wb2.close()


class TestChartCreation:
    """Test that charts can be added to workbooks."""

    def test_bar_chart_added(self, tmp_output_dir: Path) -> None:
        """A bar chart should appear in the worksheet's chart list."""
        path = tmp_output_dir / "test_chart.xlsx"
        wb = Workbook()
        ws = wb.active

        # Write data for the chart.
        ws.append(["Category", "Value"])
        ws.append(["Revenue", 10000])
        ws.append(["COGS", 6000])
        ws.append(["Gross Profit", 4000])

        # Create a bar chart.
        chart = BarChart()
        chart.title = "Income Summary"
        chart.y_axis.title = "USD ($M)"
        data = Reference(ws, min_col=2, min_row=1, max_row=4)
        categories = Reference(ws, min_col=1, min_row=2, max_row=4)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "D2")

        wb.save(str(path))

        # Verify the chart exists in the reloaded file.
        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert len(ws2._charts) == 1
        # After round-trip, chart.title is a Title object.  Extract the
        # text from its rich-text paragraph runs.
        title_obj = ws2._charts[0].title
        title_text = title_obj.tx.rich.paragraphs[0].r[0].t
        assert title_text == "Income Summary"
        wb2.close()

    def test_multiple_charts(self, tmp_output_dir: Path) -> None:
        """Multiple charts can coexist on a single sheet."""
        path = tmp_output_dir / "test_multi_chart.xlsx"
        wb = Workbook()
        ws = wb.active

        ws.append(["Quarter", "Revenue", "Expenses"])
        ws.append(["Q1", 2500, 1800])
        ws.append(["Q2", 2700, 1900])
        ws.append(["Q3", 2900, 2000])
        ws.append(["Q4", 3100, 2100])

        # Chart 1: Revenue.
        c1 = BarChart()
        c1.title = "Revenue by Quarter"
        data1 = Reference(ws, min_col=2, min_row=1, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        c1.add_data(data1, titles_from_data=True)
        c1.set_categories(cats)
        ws.add_chart(c1, "E2")

        # Chart 2: Expenses.
        c2 = BarChart()
        c2.title = "Expenses by Quarter"
        data2 = Reference(ws, min_col=3, min_row=1, max_row=5)
        c2.add_data(data2, titles_from_data=True)
        c2.set_categories(cats)
        ws.add_chart(c2, "E18")

        wb.save(str(path))

        wb2 = load_workbook(str(path))
        assert len(wb2.active._charts) == 2
        wb2.close()


class TestSaveToDisk:
    """Test workbook save and file integrity."""

    def test_save_creates_file(self, tmp_output_dir: Path) -> None:
        """Saving a workbook should create a file at the specified path."""
        path = tmp_output_dir / "test_save.xlsx"
        assert not path.exists()

        wb = Workbook()
        wb.active["A1"] = "test"
        wb.save(str(path))

        assert path.exists()
        assert path.stat().st_size > 0

    def test_saved_file_is_valid_zip(self, tmp_output_dir: Path) -> None:
        """An .xlsx file is a ZIP archive; verify it has valid ZIP structure."""
        import zipfile

        path = tmp_output_dir / "test_zip.xlsx"
        wb = Workbook()
        wb.active["A1"] = "zipcheck"
        wb.save(str(path))

        assert zipfile.is_zipfile(str(path))

    def test_save_with_multiple_sheets(self, tmp_output_dir: Path) -> None:
        """All sheets should be present in the saved file."""
        path = tmp_output_dir / "test_multi_sheet.xlsx"
        wb = Workbook()
        wb.active.title = "Assumptions"
        wb.create_sheet("DCF Model")
        wb.create_sheet("Summary")
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        assert wb2.sheetnames == ["Assumptions", "DCF Model", "Summary"]
        wb2.close()

    def test_overwrite_existing_file(self, tmp_output_dir: Path) -> None:
        """Saving to an existing path should overwrite without error."""
        path = tmp_output_dir / "test_overwrite.xlsx"

        # Create first version.
        wb1 = Workbook()
        wb1.active["A1"] = "version1"
        wb1.save(str(path))

        # Overwrite with second version.
        wb2 = Workbook()
        wb2.active["A1"] = "version2"
        wb2.save(str(path))

        # Verify the second version persisted.
        wb3 = load_workbook(str(path))
        assert wb3.active["A1"].value == "version2"
        wb3.close()
