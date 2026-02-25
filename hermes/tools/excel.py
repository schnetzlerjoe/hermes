"""Excel workbook tools for creating and manipulating financial models.

Built on openpyxl.  Provides cell-level control, formula support, formatting,
and chart embedding -- designed for building proper financial models, not just
data dumps.
"""

from __future__ import annotations

import logging
import uuid
from pathlib import Path
from typing import Any

from llama_index.core.tools import FunctionTool
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries

from hermes.config import get_config

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# In-memory workbook store
# ---------------------------------------------------------------------------

# Maps workbook IDs to open Workbook objects.  Keeps workbooks alive across
# multiple tool calls within a single agent session.
_workbooks: dict[str, Workbook] = {}


def _get_workbook(workbook_id: str) -> Workbook:
    """Retrieve an open workbook or raise a clear error."""
    if workbook_id not in _workbooks:
        raise ValueError(
            f"Workbook '{workbook_id}' not found.  Available: {list(_workbooks.keys())}"
        )
    return _workbooks[workbook_id]


# ---------------------------------------------------------------------------
# Tool functions
# ---------------------------------------------------------------------------


def excel_create_workbook(
    name: str,
    sheets: list[str] | None = None,
) -> str:
    """Create a new Excel workbook.

    The workbook is held in memory until :func:`excel_save` is called.
    If no sheet names are provided a single ``Sheet1`` is created.

    Args:
        name: Human-readable name used as the workbook identifier.
        sheets: Optional list of sheet names to create.

    Returns:
        Workbook ID string for use in subsequent operations.
    """
    wb = Workbook()
    workbook_id = f"{name}_{uuid.uuid4().hex[:8]}"

    if sheets:
        # Rename the default sheet to the first name, then add the rest.
        ws = wb.active
        ws.title = sheets[0]
        for sheet_name in sheets[1:]:
            wb.create_sheet(title=sheet_name)
    else:
        wb.active.title = "Sheet1"

    _workbooks[workbook_id] = wb
    sheet_names = [s.title for s in wb.worksheets]
    logger.info("Created workbook %s", workbook_id)
    return f"Created workbook '{workbook_id}' with sheets: {sheet_names}"


def excel_write_cells(
    workbook_id: str,
    sheet: str,
    data: dict[str, Any],
) -> str:
    """Write values to cells in a worksheet.

    Args:
        workbook_id: ID returned by :func:`excel_create_workbook`.
        sheet: Target sheet name.
        data: Mapping of cell references to values
            (e.g. ``{"A1": "Revenue", "B1": 1000000}``).

    Returns:
        Confirmation string with the number of cells written.
    """
    wb = _get_workbook(workbook_id)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found.  Available: {wb.sheetnames}")

    ws = wb[sheet]
    for cell_ref, value in data.items():
        ws[cell_ref] = value

    return f"Wrote {len(data)} cell(s) to '{sheet}'."


def excel_read_range(
    workbook_id: str,
    sheet: str,
    range_str: str,
) -> list[list]:
    """Read values from a rectangular cell range.

    Args:
        workbook_id: Workbook ID.
        sheet: Sheet name.
        range_str: Excel-style range (e.g. ``"A1:D10"``).

    Returns:
        2D list of cell values (rows of columns).
    """
    wb = _get_workbook(workbook_id)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found.  Available: {wb.sheetnames}")

    ws = wb[sheet]
    rows = []
    for row in ws[range_str]:
        rows.append([cell.value for cell in row])
    return rows


def excel_add_formula(
    workbook_id: str,
    sheet: str,
    cell: str,
    formula: str,
) -> str:
    """Insert an Excel formula at a specific cell.

    The formula string should start with ``=`` just as in Excel
    (e.g. ``"=SUM(B2:B10)"``).

    Args:
        workbook_id: Workbook ID.
        sheet: Sheet name.
        cell: Cell reference (e.g. ``"C11"``).
        formula: Excel formula string.

    Returns:
        Confirmation string.
    """
    wb = _get_workbook(workbook_id)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found.  Available: {wb.sheetnames}")

    ws = wb[sheet]
    ws[cell] = formula
    return f"Formula '{formula}' set at {sheet}!{cell}."


def excel_format_range(
    workbook_id: str,
    sheet: str,
    range_str: str,
    bold: bool = False,
    number_format: str | None = None,
    font_size: int | None = None,
    bg_color: str | None = None,
    border: bool = False,
) -> str:
    """Apply formatting to a range of cells.

    Args:
        workbook_id: Workbook ID.
        sheet: Sheet name.
        range_str: Excel-style range (e.g. ``"A1:D1"``).
        bold: Make text bold.
        number_format: Excel number format string
            (e.g. ``"#,##0"``, ``"$#,##0.00"``, ``"0.0%"``).
        font_size: Font size in points.
        bg_color: Background fill colour as a hex string without ``#``
            (e.g. ``"4472C4"``).
        border: Apply thin borders around each cell.

    Returns:
        Confirmation string.
    """
    wb = _get_workbook(workbook_id)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found.  Available: {wb.sheetnames}")

    ws = wb[sheet]

    thin_side = Side(style="thin") if border else None
    cell_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    ) if border else None

    for row in ws[range_str]:
        for cell_obj in row:
            if bold or font_size:
                existing_font = cell_obj.font
                cell_obj.font = Font(
                    bold=bold if bold else existing_font.bold,
                    size=font_size if font_size else existing_font.size,
                    name=existing_font.name,
                )
            if number_format:
                cell_obj.number_format = number_format
            if bg_color:
                cell_obj.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
            if cell_border:
                cell_obj.border = cell_border

    return f"Formatted range {range_str} on '{sheet}'."


def excel_add_chart(
    workbook_id: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    position: str,
    title: str | None = None,
) -> str:
    """Add a chart to a worksheet.

    The ``data_range`` should include column headers in the first row and
    category labels in the first column.

    Args:
        workbook_id: Workbook ID.
        sheet: Sheet name.
        chart_type: One of ``"line"``, ``"bar"``, ``"pie"``, ``"scatter"``.
        data_range: Range containing chart data (e.g. ``"A1:D10"``).
        position: Cell where the chart's top-left corner is anchored
            (e.g. ``"F2"``).
        title: Optional chart title.

    Returns:
        Confirmation string.
    """
    wb = _get_workbook(workbook_id)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found.  Available: {wb.sheetnames}")

    ws = wb[sheet]

    chart_classes = {
        "line": LineChart,
        "bar": BarChart,
        "pie": PieChart,
        "scatter": ScatterChart,
    }
    chart_cls = chart_classes.get(chart_type.lower())
    if chart_cls is None:
        raise ValueError(
            f"Unsupported chart type '{chart_type}'.  "
            f"Choose from: {list(chart_classes.keys())}"
        )

    chart = chart_cls()
    if title:
        chart.title = title

    min_col, min_row, max_col, max_row = range_boundaries(data_range)

    # Categories from the first column (excluding header).
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)

    # Each subsequent column is a data series.
    for col_idx in range(min_col + 1, max_col + 1):
        values = Reference(ws, min_col=col_idx, min_row=min_row, max_row=max_row)
        chart.add_data(values, titles_from_data=True)

    chart.set_categories(cats)
    chart.width = 18
    chart.height = 12

    ws.add_chart(chart, position)
    return f"Added {chart_type} chart at {position} on '{sheet}'."


def excel_add_sheet(workbook_id: str, sheet_name: str) -> str:
    """Add a new worksheet to an existing workbook.

    Args:
        workbook_id: Workbook ID.
        sheet_name: Name for the new sheet.

    Returns:
        Confirmation string.
    """
    wb = _get_workbook(workbook_id)
    if sheet_name in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' already exists in workbook '{workbook_id}'.")
    wb.create_sheet(title=sheet_name)
    return f"Added sheet '{sheet_name}' to workbook '{workbook_id}'."


def excel_save(
    workbook_id: str,
    filename: str | None = None,
) -> str:
    """Save workbook to disk as an ``.xlsx`` file.

    Saves to the configured output directory.  The workbook remains open
    in memory for further modifications after saving.

    Args:
        workbook_id: Workbook ID.
        filename: Optional filename (without path).  Defaults to
            ``"{workbook_id}.xlsx"``.

    Returns:
        Absolute path to the saved file.
    """
    wb = _get_workbook(workbook_id)
    cfg = get_config()

    output_dir = Path(cfg.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if filename is None:
        filename = f"{workbook_id}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    filepath = output_dir / filename
    wb.save(str(filepath))
    logger.info("Saved workbook to %s", filepath.resolve())
    return str(filepath.resolve())


# ---------------------------------------------------------------------------
# Tool registration
# ---------------------------------------------------------------------------


def create_tools() -> list[FunctionTool]:
    """Create LlamaIndex FunctionTool instances for all Excel tools."""
    return [
        FunctionTool.from_defaults(
            fn=excel_create_workbook,
            name="excel_create_workbook",
            description=(
                "Create a new Excel workbook with optional named sheets. "
                "Returns a workbook ID for subsequent operations."
            ),
        ),
        FunctionTool.from_defaults(
            fn=excel_write_cells,
            name="excel_write_cells",
            description=(
                "Write values to specific cells. Provide a dict mapping cell "
                "references to values (e.g. {'A1': 'Revenue', 'B1': 1000000})."
            ),
        ),
        FunctionTool.from_defaults(
            fn=excel_read_range,
            name="excel_read_range",
            description=(
                "Read values from a cell range (e.g. 'A1:D10'). Returns a "
                "2D list of cell values."
            ),
        ),
        FunctionTool.from_defaults(
            fn=excel_add_formula,
            name="excel_add_formula",
            description=(
                "Insert an Excel formula at a specific cell. The formula "
                "string should start with '=' (e.g. '=SUM(B2:B10)')."
            ),
        ),
        FunctionTool.from_defaults(
            fn=excel_format_range,
            name="excel_format_range",
            description=(
                "Apply formatting to a range of cells: bold, number format "
                "(e.g. '#,##0', '$#,##0.00', '0.0%'), font size, background "
                "colour, and borders."
            ),
        ),
        FunctionTool.from_defaults(
            fn=excel_add_chart,
            name="excel_add_chart",
            description=(
                "Add a chart (line, bar, pie, scatter) to a sheet. Specify the "
                "data range, chart position, and optional title."
            ),
        ),
        FunctionTool.from_defaults(
            fn=excel_add_sheet,
            name="excel_add_sheet",
            description="Add a new sheet to an existing workbook.",
        ),
        FunctionTool.from_defaults(
            fn=excel_save,
            name="excel_save",
            description=(
                "Save the workbook to disk as .xlsx. Returns the absolute "
                "file path."
            ),
        ),
    ]
